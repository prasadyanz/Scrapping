import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


s = Service(r'C:\Users\Prasad\Desktop\chromedriver.exe')
driver = webdriver.Chrome(service = s)

################################################################################################
location = r"C:\Users\Prasad\My Drive\My_Finance_Datas\Stock_data\Individual_finance_data_BSE/"

all_excels=[]
for file in os.listdir(location):
    if '.xlsx' in file:
        all_excels.append(file)

##################################################################################################
def get_fin_from_BSE(bse_path):

    driver.get(bse_path)

    # elem = driver.find_element(by=By.ID, value="ContentPlaceHolder1_tdData")
    # elem = driver.find_element(by = By.CLASS_NAME, value= "TTRow")

    elems_site_1 = driver.find_elements(by = By.XPATH, value = "//a[@href]")

    list_of_link_for_AN = []

    # get the list of all the links in this sight
    for elem_site_1 in elems_site_1:
        # print(elem_site_1.get_attribute("href"))
        link_temp = elem_site_1.get_attribute("href")
        link = link_temp.split(".")

        # finding the link based on .50& in the link
        if link[len(link)-1][:2] == "50":
            list_of_link_for_AN.append(link_temp)
            # print(link_temp)

    def add_data_to_list(link_list):
        driver.get(link_list)
        data_site_2 = driver.find_elements(by=By.XPATH, value="*")
        temp_list_oneyear = []
        def consolidated():
            global no
            no = 0
            check = data_site_2[0].text
            check = check.split("\n")
            temp_list = []
            for y in range(0, len(check)):
                if check[y] == "Corporate Filings":
                    no = 1
                elif check[y] == "Notes":
                    no = 2
                elif no == 1:
                    temp_list.append(check[y])
                    # print(check[y])
                else:
                    pass
            return temp_list
        def standalone():
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
                (By.XPATH, "//div[@id='divmain']//a[@id='ContentPlaceHolder1_lnkDetailed']"))).click()
            data_site_3 = driver.find_elements(by=By.XPATH, value="*")
            global no
            no = 0
            check = data_site_3[0].text
            check = check.split("\n")
            temp_list = []
            for y in range(0, len(check)):
                if check[y] == "Corporate Filings":
                    no = 1
                elif check[y] == "Notes":
                    no = 2
                elif no == 1:
                    temp_list.append(check[y])
                    # print(check[y])
                else:
                    pass
            return temp_list

        temp_list_oneyear.append(consolidated())
        temp_list_oneyear.append(standalone())
        return temp_list_oneyear

    # open each years Annual financals
    bse_fin_data = []
    for x in list_of_link_for_AN:
        # print(x)
        # print(add_data_to_list(x))
    # print(list_of_link_for_AN[0])
    # print(add_data_to_list(list_of_link_for_AN[0])[0])
    # print(add_data_to_list(list_of_link_for_AN[0])[1])
        bse_fin_data.append(add_data_to_list(x))

    driver.close()
    return bse_fin_data

def write_in_excel(bse_data, codename):
    companyname = (bse_data[0][0][3].split(":")[2])
    excelname = companyname + "_" + codename + ".xlsx"

    excel_path = location + excelname
    if excel_path in all_excels:
        print(excelname + "   is available ")
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Consolidated Results"
        wb.create_sheet("Detailed Results")
        wb.save(excel_path)
        print(excelname + "   Saved ")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    # print(excel_path)
    for x in range(0, len(bse_data)):
        for y in bse_data[x]:
            print(y)
            for z in range(2, len(y)):
                sheet = wb[y[1]]
                sheet.cell(row=z, column=x + 1).value = y[z]
    wb.save(excel_path)

def sort_data(bse_data):
    main_heading = []
    temp_data = []

    def mysplit(s):
        head = s.rstrip('-,.0123456789')
        tail = s[len(head):]
        return (head, tail)

    global status
    for x in range(0, len(bse_data)):
        for y in bse_data[x]:
            status = "Pass"
            for z in y:
                if z.split(" ")[0]=="Note":
                    status = "Skip"
                else:
                    pass

            if status == "Pass":
                temp_data_1 = [mysplit(s) for s in y]
                temp_data.append(temp_data_1)
                for dat in temp_data_1:
                    heading_temp = dat[0]
                    if heading_temp in main_heading:
                        pass
                    else:
                        main_heading.append(heading_temp)



    # print(main_heading)
    global availability
    for y in main_heading:
        availability = "False"
        temp_1 = [y]
        for x in temp_data:
            # print(x)
            for z in x :
                if y == str(z[0]):
                    temp_1.append(z[1])
                    availability = "True"
                else:
                    pass
            if availability == "False":
                temp_1.append("NA")
        print(len(temp_data),len(temp_1), temp_1)





if __name__ == '__main__':
    codename = "500209"
    path = 'https://www.bseindia.com/corporates/Comp_Results.aspx?Code='+codename
    # get_fin_from_BSE(path)
    bse_data = get_fin_from_BSE(path)
    sort_data(bse_data)
    # write_in_excel(bse_data, codename)
