import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


s = Service(r'C:\Users\Prasad\Desktop\chromedriver.exe')
driver = webdriver.Chrome(service = s)

def get_fin_from_BSE(bse_path):

    driver.get(bse_path)

    # elem = driver.find_element(by=By.ID, value="ContentPlaceHolder1_tdData")
    # elem = driver.find_element(by = By.CLASS_NAME, value= "TTRow")

    elems_site_1 = driver.find_elements(by = By.XPATH, value = "//a[@href]")

    list_of_link_for_AN = []

    # get the list of all the links in this sight
    for elem_site_1 in elems_site_1:
        # print(elem_site_1.get_attribute("href"))
        link_temp = elem_site_1.get_attribute("href")
        link = link_temp.split(".")

        # finding the link based on .50& in the link
        if link[len(link)-1][:2] == "50":
            list_of_link_for_AN.append(link_temp)
            # print(link_temp)

    def add_data_to_list(link_list):
        driver.get(link_list)
        data_site_2 = driver.find_elements(by=By.XPATH, value="*")
        temp_list_oneyear = []
        def consolidated():
            global no
            no = 0
            check = data_site_2[0].text
            check = check.split("\n")
            temp_list = []
            for y in range(0, len(check)):
                if check[y] == "Corporate Filings":
                    no = 1
                elif check[y] == "Notes":
                    no = 2
                elif no == 1:
                    temp_list.append(check[y])
                    # print(check[y])
                else:
                    pass
            return temp_list
        def standalone():
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
                (By.XPATH, "//div[@id='divmain']//a[@id='ContentPlaceHolder1_lnkDetailed']"))).click()
            data_site_3 = driver.find_elements(by=By.XPATH, value="*")
            global no
            no = 0
            check = data_site_3[0].text
            check = check.split("\n")
            temp_list = []
            for y in range(0, len(check)):
                if check[y] == "Corporate Filings":
                    no = 1
                elif check[y] == "Notes":
                    no = 2
                elif no == 1:
                    temp_list.append(check[y])
                    # print(check[y])
                else:
                    pass
            return temp_list

        temp_list_oneyear.append(consolidated())
        temp_list_oneyear.append(standalone())
        return temp_list_oneyear

    # open each years Annual financals
    bse_fin_data = []
    for x in list_of_link_for_AN:
        # print(x)
        # print(add_data_to_list(x))
    # print(list_of_link_for_AN[0])
    # print(add_data_to_list(list_of_link_for_AN[0])[0])
    # print(add_data_to_list(list_of_link_for_AN[0])[1])
        bse_fin_data.append(add_data_to_list(x))

    driver.close()
    return bse_fin_data

if __name__ == '__main__':
    codename = "500209"
    location = r"C:\Users\Prasad\My Drive\My_Finance_Datas\Stock_data\Individual_finance_data_BSE/"


    path = 'https://www.bseindia.com/corporates/Comp_Results.aspx?Code='+codename
    # get_fin_from_BSE(path)
    bse_data = get_fin_from_BSE(path)
    companyname = (bse_data[0][0][3].split(":")[2])
    excelname = companyname+"_"+codename+".xlsx"

    excel_path = location + excelname
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Consolidated Results"
    wb.create_sheet("Detailed Results")
    # wb.save(excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(excel_path)
    for x in range(0, len(bse_data)):
        for y in bse_data[x]:
            print(y)
            for z in range(2, len(y)):
                sheet = wb[y[1]]
                sheet.cell(row = z, column = x+1).value = y[z]

wb.save(excel_path)